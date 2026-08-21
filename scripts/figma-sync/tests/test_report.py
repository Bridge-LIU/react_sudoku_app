"""v11.3 (figma-sync 専用 7 sheet, code from scratch) テスト。

シート構成:
  01_表紙サマリ / 02_変更点サマリ / 03_UI Before-After /
  04_コード変更 / 05_全システムテスト / 06_異常系＆セキュリティ / 07_承認履歴

v11.3 でテンプレファイル load 型から code from scratch に転換したため、
template ファイル存在チェックは廃止。fill_report は Workbook を code で組み立てる。
"""
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from scripts.lib.excel_fill import (  # noqa: E402
    fill_report, ExcelFillContext, SHEET_ORDER,
    _fmt_variable, _rgba_to_hex, normalize_variables_payload,
)


EXPECTED_SHEETS = [
    '01_表紙サマリ',
    '02_変更点サマリ',
    '03_UI Before-After',
    '04_コード変更',
    '05_全システムテスト',
    '06_異常系＆セキュリティ',
    '07_承認履歴',
]


def _base_ctx(tmp_path, **overrides):
    kwargs = dict(
        template_path='',  # v11.3: 未使用
        output_path=str(tmp_path / 'report.xlsx'),
        run_ts='2026-08-21T10-00-00-000Z',
        test_results={
            'unit': {'exitCode': 0, 'numTotal': 100, 'numPassed': 100, 'numFailed': 0, 'duration': 6.2},
            'e2e': {'exitCode': 0, 'numTotal': 5, 'numPassed': 5, 'numFailed': 0, 'duration': 42.1},
            'audit': {'json': {'metadata': {'vulnerabilities': {
                'critical': 0, 'high': 0, 'moderate': 0, 'low': 0, 'info': 0, 'total': 0}}}},
            'coverage': {'total': {'lines': {'pct': 92.3}}},
        },
        design_changes=[],
        config={
            'figmaFileKey': 'testKey123',
            'figmaFileUrl': 'https://figma.com/design/testKey123/',
            'frames': [
                {'nodeId': '11:1896', 'file': 'src/app/index.tsx', 'component': 'Home'},
                {'nodeId': '18:6', 'file': 'src/app/play/[difficulty].tsx', 'component': 'Play'},
            ],
        },
        state_before={'last_version_id': 'OLD_ID', 'last_run_at': '2026-08-20T00:00:00Z'},
        state_after={'last_version_id': 'NEW_ID', 'last_run_at': '2026-08-21T10:00:00Z'},
        apply_result={'changedFiles': ['src/app/play/[difficulty].tsx'], 'errors': []},
        diff_summary={'nodeDiffCount': 1, 'warnings': []},
        git_info={
            'commit_sha': 'abcdef1234567890',
            'commit_msg': 'feat(figma-sync): apply nodeDiff',
            'changed_files': ['src/app/play/[difficulty].tsx'],
            'pr_url': None,
        },
        phase_results={
            'phase_1a': 'CHANGED',
            'phase_1b': '1 NodeDiff',
            'phase_1b_fallback': 'SKIP',
            'phase_3': '1 files apply, 0 errors',
            'phase_4': '3 PNG',
            'phase_5': 'unit=PASS e2e=PASS',
            'phase_7': 'APPROVED',
        },
        final_status='APPROVED',
    )
    kwargs.update(overrides)
    return ExcelFillContext(**kwargs)


# ─────────────────────────────────────────────────
# シート構成テスト
# ─────────────────────────────────────────────────
def test_sheet_order_constant_is_7_sheets():
    """SHEET_ORDER 定数が正しく 7 シート"""
    assert SHEET_ORDER == EXPECTED_SHEETS


def test_fill_report_creates_file_with_7_sheets(tmp_path):
    """fill_report 実行後、出力に 7 シートが正しい順で並ぶ"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    assert Path(ctx.output_path).exists()
    wb = load_workbook(str(ctx.output_path))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_fill_report_no_template_file_needed(tmp_path):
    """template_path が空文字でも fill_report は動く (v11.3 で不要化)"""
    ctx = _base_ctx(tmp_path)
    ctx.template_path = ''
    fill_report(ctx)
    assert Path(ctx.output_path).exists()


# ─────────────────────────────────────────────────
# Sheet 01: 表紙サマリ
# ─────────────────────────────────────────────────
def test_sheet_01_cover_has_title_banner(tmp_path):
    """A1 = テスト実施報告書、A2 = react_sudoku_app サブタイトル"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    assert ws.cell(1, 1).value == 'テスト実施報告書'
    assert 'react_sudoku_app' in str(ws.cell(2, 1).value or '')
    assert 'v11.3' in str(ws.cell(2, 1).value or '')


def test_sheet_01_cover_has_project_meta(tmp_path):
    """メタ表にプロジェクト名/final status/コミット sha が入る"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    # メタ表 (row 4-13) から探す
    meta_values = {}
    for r in range(4, 14):
        k = str(ws.cell(r, 1).value or '')
        v = str(ws.cell(r, 4).value or '')
        if k:
            meta_values[k] = v
    assert meta_values.get('プロジェクト名') == 'react_sudoku_app'
    assert meta_values.get('final status') == 'APPROVED'
    assert 'abcdef12' in meta_values.get('コミット', '')


def test_sheet_01_toc_has_hyperlinks_to_all_other_sheets(tmp_path):
    """目次に他 6 シートへのハイパーリンクが並ぶ"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    links = []
    for row in ws.iter_rows():
        for c in row:
            if c.hyperlink:
                links.append((c.value, c.hyperlink.location or c.hyperlink.target))
    assert len(links) == 6
    target_sheets = {v for v, _ in links}
    for name in EXPECTED_SHEETS[1:]:
        assert name in target_sheets, f'目次リンク欠落: {name}'


def test_sheet_01_fallback_banner_shows_when_triggered(tmp_path):
    """fallback_triggered=True の run で 01_表紙サマリ A3 に赤字 banner"""
    fi = {'reason': 'DIFF_EMPTY', 'nodeDiffCount': 1, 'fallbackStatus': 'CHANGED'}
    ctx = _base_ctx(tmp_path, fallback_triggered=True, fallback_info=fi)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    banner = str(ws.cell(3, 1).value or '')
    assert 'FALLBACK' in banner
    assert 'DIFF_EMPTY' in banner


def test_sheet_01_fallback_banner_absent_when_not_triggered(tmp_path):
    """通常経路では banner が空"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    assert ws.cell(3, 1).value in (None, '')


def test_sheet_01_result_panel_shows_ok_ng_warn(tmp_path):
    """結果パネルに OK/NG/WARN ラベルと総合判定"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['01_表紙サマリ']
    # パネル label 行を探す
    labels = []
    for row in ws.iter_rows():
        for c in row:
            v = str(c.value or '')
            if v in ('OK', 'NG', 'WARN', '総合判定'):
                labels.append(v)
    assert 'OK' in labels
    assert 'NG' in labels
    assert 'WARN' in labels
    assert '総合判定' in labels


# ─────────────────────────────────────────────────
# Sheet 02: 変更点サマリ
# ─────────────────────────────────────────────────
def test_sheet_02_change_summary_empty_dc_shows_placeholder(tmp_path):
    """design_changes 空なら「-」種別の placeholder 行"""
    ctx = _base_ctx(tmp_path, design_changes=[])
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['02_変更点サマリ']
    # header 行 4、データ 5-
    assert '変更検出なし' in str(ws.cell(5, 4).value or '')


def test_sheet_02_change_summary_lists_variable_and_screen(tmp_path):
    """bindingChanges と非 binding の両方が Variable/Screen として展開"""
    dc = [
        {
            'component': 'Play', 'file': 'src/app/play/[difficulty].tsx', 'node_id': '18:6',
            'changes': [{'bindingChanges': [
                {'node_id': '147:2', 'node_name': 'div.x', 'property': 'fills[0]',
                 'from_variable_id': None, 'to_variable_id': 'VariableID:3:4',
                 'change_kind': 'added'},
            ]}],
        },
        {
            'component': 'Home', 'file': 'src/app/index.tsx', 'node_id': '11:1896',
            'changes': [{'change_kind': 'modified'}],
        },
    ]
    ctx = _base_ctx(tmp_path, design_changes=dc)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['02_変更点サマリ']
    kinds = [str(ws.cell(r, 2).value or '') for r in range(5, 8)]
    assert 'Variable' in kinds
    assert 'Screen' in kinds


# ─────────────────────────────────────────────────
# Sheet 03: UI Before-After
# ─────────────────────────────────────────────────
def test_sheet_03_ui_empty_shows_placeholder(tmp_path):
    """design_changes 空なら placeholder 表示"""
    ctx = _base_ctx(tmp_path, design_changes=[])
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['03_UI Before-After']
    assert 'design_changes 空' in str(ws.cell(5, 1).value or '') or \
           '変更なし' in str(ws.cell(5, 1).value or '')


def test_sheet_03_ui_row_per_design_change(tmp_path):
    """design_change 1 件 = 1 行 (画面名/Before/After/差分説明)"""
    dc = [{
        'component': 'Play', 'file': 'src/app/play/[difficulty].tsx', 'node_id': '18:6',
        'changes': [{'bindingChanges': [
            {'node_id': '147:2', 'node_name': 'div.x', 'property': 'fills[0]',
             'from_variable_id': None, 'to_variable_id': 'VariableID:3:4',
             'change_kind': 'added'},
        ]}],
        'before_png': None, 'after_png': None, 'diff_pixels': 1234,
    }]
    ctx = _base_ctx(tmp_path, design_changes=dc)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['03_UI Before-After']
    # 行 5: idx=1, 画面=Play (18:6)
    assert ws.cell(5, 1).value == 1
    assert 'Play' in str(ws.cell(5, 2).value or '')
    # Before/After は画像なしで label 出力
    assert 'BEFORE' in str(ws.cell(5, 3).value or '')
    assert 'AFTER' in str(ws.cell(5, 4).value or '')
    # 差分説明に binding 数 + diff_pixels
    e = str(ws.cell(5, 5).value or '')
    assert '1 binding' in e
    assert '1234' in e


# ─────────────────────────────────────────────────
# Sheet 04: コード変更
# ─────────────────────────────────────────────────
def test_sheet_04_code_changes_lists_apply_files(tmp_path):
    """apply_result.changedFiles が行 5- に並ぶ"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    # 行 5: idx=1, ファイル, sha, msg
    assert ws.cell(5, 1).value == 1
    assert ws.cell(5, 2).value == 'src/app/play/[difficulty].tsx'
    assert 'abcdef12' in str(ws.cell(5, 3).value or '')


def test_sheet_04_code_changes_fallback_route_column(tmp_path):
    """fallback 発動時、検出経路列が fallback 表記になる"""
    fi = {'reason': 'DIFF_EMPTY', 'nodeDiffCount': 1, 'fallbackStatus': 'CHANGED'}
    ctx = _base_ctx(tmp_path, fallback_triggered=True, fallback_info=fi)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    # 行 4 F列 = header「検出経路」
    assert str(ws.cell(4, 6).value or '') == '検出経路'
    # 行 5 F列 = fallback ラベル
    assert 'fallback' in str(ws.cell(5, 6).value or '').lower()


def test_sheet_04_binding_section_expanded(tmp_path):
    """◆ Figma 変更詳細 section と bindingChanges 展開行"""
    dc = [{
        'component': 'Play', 'file': 'src/app/play/[difficulty].tsx', 'node_id': '18:6',
        'changes': [{'bindingChanges': [
            {'node_id': '147:2', 'node_name': 'div.x', 'property': 'minWidth',
             'from_variable_id': None, 'to_variable_id': 'VariableID:11:2315',
             'change_kind': 'added'},
        ]}],
    }]
    ctx = _base_ctx(tmp_path, design_changes=dc)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    # section header を含む行を探す
    header_row = None
    for r in range(1, ws.max_row + 1):
        if 'Figma 変更詳細' in str(ws.cell(r, 1).value or ''):
            header_row = r
            break
    assert header_row is not None
    # header 行の次が binding table header、その次から data
    tbl_header = header_row + 1
    assert str(ws.cell(tbl_header, 1).value) == '#'
    assert str(ws.cell(tbl_header, 4).value) == 'property'
    assert str(ws.cell(tbl_header, 6).value) == '種類'
    # data 行
    data_row = tbl_header + 1
    assert ws.cell(data_row, 1).value == 1
    assert 'Play' in str(ws.cell(data_row, 2).value or '')
    assert str(ws.cell(data_row, 4).value) == 'minWidth'
    assert str(ws.cell(data_row, 6).value) == 'added'


def test_sheet_04_binding_section_empty_shows_placeholder(tmp_path):
    """bindingChanges なし → placeholder"""
    ctx = _base_ctx(tmp_path, design_changes=[])
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    # placeholder テキストを含む row を検索
    found = False
    for r in range(1, ws.max_row + 1):
        if 'bindingChanges なし' in str(ws.cell(r, 1).value or ''):
            found = True
            break
    assert found


def test_sheet_04_binding_uses_variables_map(tmp_path):
    """variables_map ありなら from/to 列に name (value) 表示"""
    dc = [{
        'component': 'Play', 'file': 'src/app/play/[difficulty].tsx', 'node_id': '18:6',
        'changes': [{'bindingChanges': [
            {'node_id': '147:2', 'node_name': 'div.x', 'property': 'fills[0]',
             'from_variable_id': None, 'to_variable_id': 'VariableID:3:4',
             'change_kind': 'added'},
        ]}],
    }]
    vmap = {'VariableID:3:4': {'name': 'color/primary', 'resolved_type': 'COLOR',
                                'value': {'r': 0.117647, 'g': 0.564706, 'b': 1.0, 'a': 1.0}}}
    ctx = _base_ctx(tmp_path, design_changes=dc, variables_map=vmap)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    # binding data 行を探す (from = '(なし)' の行)
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 5).value or '').startswith('(なし)'):
            to_cell = str(ws.cell(r, 5).value or '')
            assert 'color/primary' in to_cell
            assert '#1E90FF' in to_cell
            assert 'VariableID:3:4' not in to_cell
            return
    pytest.fail('from=(なし) の binding 行が見つからず')


def test_sheet_04_binding_without_variables_map_shows_unresolved(tmp_path):
    """variables_map 空なら raw ID + (未解決) に degradation"""
    dc = [{
        'component': 'Play', 'file': 'src/app/play/[difficulty].tsx', 'node_id': '18:6',
        'changes': [{'bindingChanges': [
            {'node_id': '147:2', 'node_name': 'div.x', 'property': 'fills[0]',
             'from_variable_id': None, 'to_variable_id': 'VariableID:3:4',
             'change_kind': 'added'},
        ]}],
    }]
    ctx = _base_ctx(tmp_path, design_changes=dc)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['04_コード変更']
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(r, 5).value or '')
        if 'VariableID:3:4' in v:
            assert '(未解決)' in v
            return
    pytest.fail('VariableID:3:4 (未解決) を含む行なし')


# ─────────────────────────────────────────────────
# Sheet 05: 全システムテスト
# ─────────────────────────────────────────────────
def test_sheet_05_full_tests_has_vitest_playwright_rows(tmp_path):
    """vitest / Playwright / a11y / snapshot の 4 種別行"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['05_全システムテスト']
    tool_col_values = [str(ws.cell(r, 3).value or '') for r in range(5, 9)]
    assert 'Vitest' in tool_col_values
    assert 'Playwright' in tool_col_values
    assert 'axe-core' in tool_col_values
    assert 'Jest' in tool_col_values


def test_sheet_05_full_tests_reflects_pass_counts(tmp_path):
    """test_results.unit.numPassed が反映される"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['05_全システムテスト']
    # Vitest 行を探して PASS 数チェック
    for r in range(5, 9):
        if str(ws.cell(r, 3).value or '') == 'Vitest':
            assert ws.cell(r, 4).value == 100  # numTotal
            assert ws.cell(r, 5).value == 100  # numPassed
            assert ws.cell(r, 6).value == 0    # numFailed
            return
    pytest.fail('Vitest 行なし')


def test_sheet_05_npm_audit_section(tmp_path):
    """npm audit セクションで severity 別カウント"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['05_全システムテスト']
    # 「◆ npm audit」を含む row 以降を探す
    for r in range(1, ws.max_row + 1):
        if 'npm audit' in str(ws.cell(r, 1).value or ''):
            # 次の行から critical/high/... 順に並ぶ
            severities = [str(ws.cell(r + 1 + i, 1).value or '') for i in range(6)]
            assert 'critical' in severities
            assert 'total' in severities
            return
    pytest.fail('npm audit section なし')


# ─────────────────────────────────────────────────
# Sheet 06: 異常系＆セキュリティ
# ─────────────────────────────────────────────────
def test_sheet_06_abnormal_sec_has_fallback_row(tmp_path):
    """fallback 発動時に該当行が WARN"""
    fi = {'reason': 'DIFF_EMPTY', 'nodeDiffCount': 2, 'fallbackStatus': 'CHANGED'}
    ctx = _base_ctx(tmp_path, fallback_triggered=True, fallback_info=fi)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['06_異常系＆セキュリティ']
    for r in range(5, ws.max_row + 1):
        case = str(ws.cell(r, 3).value or '')
        if 'Fallback' in case:
            assert str(ws.cell(r, 4).value) == 'WARN'
            return
    pytest.fail('Fallback 行なし')


def test_sheet_06_abnormal_sec_apply_error_row(tmp_path):
    """apply errors ありなら apply エラー行が FAIL"""
    ctx = _base_ctx(tmp_path, apply_result={
        'changedFiles': ['x.tsx'], 'errors': ['err1']
    })
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['06_異常系＆セキュリティ']
    for r in range(5, ws.max_row + 1):
        case = str(ws.cell(r, 3).value or '')
        if 'apply エラー' in case:
            assert str(ws.cell(r, 4).value) == 'FAIL'
            return
    pytest.fail('apply エラー行なし')


def test_sheet_06_audit_row_pass_when_zero_vuln(tmp_path):
    """npm audit total=0 なら PASS"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['06_異常系＆セキュリティ']
    for r in range(5, ws.max_row + 1):
        case = str(ws.cell(r, 3).value or '')
        if 'npm audit' in case:
            assert str(ws.cell(r, 4).value) == 'PASS'
            return
    pytest.fail('npm audit 行なし')


# ─────────────────────────────────────────────────
# Sheet 07: 承認履歴
# ─────────────────────────────────────────────────
def test_sheet_07_approvals_has_3_rows(tmp_path):
    """AI1 / AI2 / 人間 の 3 行"""
    ctx = _base_ctx(tmp_path)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['07_承認履歴']
    stages = [str(ws.cell(r, 2).value or '') for r in range(5, 8)]
    assert 'AI 1段目' in stages
    assert 'AI 2段目' in stages
    assert '人間' in stages


def test_sheet_07_human_verdict_shows_approved_marker_when_approved(tmp_path):
    """final_status=APPROVED なら人間欄に ☑適用可 マーカー"""
    ctx = _base_ctx(tmp_path, final_status='APPROVED')
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['07_承認履歴']
    # 人間行を探す
    for r in range(5, 8):
        if str(ws.cell(r, 2).value or '') == '人間':
            assert '適用可' in str(ws.cell(r, 5).value or '')
            return
    pytest.fail('人間行なし')


def test_sheet_07_human_verdict_shows_rejected_marker_when_rejected(tmp_path):
    """final_status=REJECTED なら人間欄に ☑否 マーカー"""
    ctx = _base_ctx(tmp_path, final_status='REJECTED')
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    ws = wb['07_承認履歴']
    for r in range(5, 8):
        if str(ws.cell(r, 2).value or '') == '人間':
            assert '否' in str(ws.cell(r, 5).value or '')
            return
    pytest.fail('人間行なし')


# ─────────────────────────────────────────────────
# 統合: fallback 全体連動
# ─────────────────────────────────────────────────
def test_fallback_triggered_reflects_on_cover_and_sheets(tmp_path):
    """fallback 発動時、01 表紙 banner + 04 検出経路列 + 06 異常系行 が連動"""
    fi = {'reason': 'DIFF_EMPTY', 'nodeDiffCount': 2, 'fallbackStatus': 'CHANGED'}
    ctx = _base_ctx(tmp_path, fallback_triggered=True, fallback_info=fi)
    fill_report(ctx)
    wb = load_workbook(str(ctx.output_path))
    # 01 banner
    assert 'FALLBACK' in str(wb['01_表紙サマリ'].cell(3, 1).value or '')
    # 04 route
    assert 'fallback' in str(wb['04_コード変更'].cell(5, 6).value or '').lower()
    # 06 fallback WARN 行
    ws06 = wb['06_異常系＆セキュリティ']
    fallback_warned = any(
        'Fallback' in str(ws06.cell(r, 3).value or '') and str(ws06.cell(r, 4).value) == 'WARN'
        for r in range(5, ws06.max_row + 1)
    )
    assert fallback_warned


# ─────────────────────────────────────────────────
# v11.2 helper (variables_map / _fmt_variable / normalize_variables_payload)
# ─────────────────────────────────────────────────
def test_fmt_variable_none_returns_nashi():
    assert _fmt_variable(None, {}) == '(なし)'
    assert _fmt_variable(None, {'VariableID:1:1': {'name': 'x', 'value': 1}}) == '(なし)'


def test_fmt_variable_empty_map_returns_raw_with_unresolved():
    assert _fmt_variable('VariableID:11:2315', {}) == 'VariableID:11:2315 (未解決)'
    assert _fmt_variable('VariableID:11:2315', None) == 'VariableID:11:2315 (未解決)'


def test_fmt_variable_resolves_name_and_value_float():
    m = {'VariableID:11:2315': {'name': 'spacing/lg', 'resolved_type': 'FLOAT', 'value': 24.0}}
    assert _fmt_variable('VariableID:11:2315', m) == 'spacing/lg (24)'


def test_fmt_variable_resolves_color_to_hex():
    m = {'VariableID:3:4': {'name': 'color/primary', 'resolved_type': 'COLOR',
                             'value': {'r': 0.117647, 'g': 0.564706, 'b': 1.0, 'a': 1.0}}}
    assert _fmt_variable('VariableID:3:4', m) == 'color/primary (#1E90FF)'


def test_fmt_variable_resolves_color_with_alpha():
    m = {'VariableID:3:5': {'name': 'color/overlay', 'resolved_type': 'COLOR',
                             'value': {'r': 0, 'g': 0, 'b': 0, 'a': 0.5}}}
    assert _fmt_variable('VariableID:3:5', m) == 'color/overlay (#00000080)'


def test_fmt_variable_boolean_and_string():
    m = {
        'VariableID:9:1': {'name': 'flag/dark', 'resolved_type': 'BOOLEAN', 'value': True},
        'VariableID:9:2': {'name': 'text/label', 'resolved_type': 'STRING', 'value': 'Hello'},
    }
    assert _fmt_variable('VariableID:9:1', m) == 'flag/dark (true)'
    assert _fmt_variable('VariableID:9:2', m) == 'text/label ("Hello")'


def test_fmt_variable_unknown_type_falls_back_to_str():
    m = {'VariableID:5:5': {'name': 'x/y', 'resolved_type': 'UNKNOWN', 'value': 42}}
    assert _fmt_variable('VariableID:5:5', m) == 'x/y (42)'


def test_rgba_to_hex_malformed_returns_str():
    out = _rgba_to_hex({'x': 'y'})
    assert isinstance(out, str)


def test_normalize_variables_payload_rest_shape():
    raw = {
        'meta': {
            'variables': {
                'VariableID:11:2315': {
                    'id': 'VariableID:11:2315', 'name': 'spacing/lg',
                    'resolvedType': 'FLOAT', 'valuesByMode': {'1:0': 24},
                },
                'VariableID:3:4': {
                    'id': 'VariableID:3:4', 'name': 'color/primary',
                    'resolvedType': 'COLOR',
                    'valuesByMode': {'1:0': {'r': 0.117647, 'g': 0.564706, 'b': 1.0, 'a': 1.0}},
                },
            }
        }
    }
    m = normalize_variables_payload(raw)
    assert 'VariableID:11:2315' in m
    assert m['VariableID:11:2315']['name'] == 'spacing/lg'
    assert m['VariableID:11:2315']['resolved_type'] == 'FLOAT'
    assert m['VariableID:11:2315']['value'] == 24
    assert m['VariableID:3:4']['value']['g'] > 0.5


def test_normalize_variables_payload_mcp_wrapped():
    raw = {'variables': {'VariableID:1:1': {'name': 'a', 'resolvedType': 'FLOAT',
                                              'valuesByMode': {'0:0': 8}}}}
    m = normalize_variables_payload(raw)
    assert m['VariableID:1:1']['value'] == 8


def test_normalize_variables_payload_flat_idempotent():
    raw = {'VariableID:1:1': {'name': 'a', 'resolved_type': 'FLOAT', 'value': 8}}
    m = normalize_variables_payload(raw)
    assert m['VariableID:1:1']['value'] == 8


def test_normalize_variables_payload_malformed_returns_empty():
    assert normalize_variables_payload(None) == {}
    assert normalize_variables_payload('') == {}
    assert normalize_variables_payload(42) == {}
    assert normalize_variables_payload({'meta': 'bogus'}) == {}
    m = normalize_variables_payload({'VariableID:x': 'not-a-dict'})
    assert m == {}


def test_normalize_variables_payload_alias_value():
    raw = {'meta': {'variables': {'VariableID:5:5': {
        'name': 'alias/x', 'resolvedType': 'FLOAT',
        'valuesByMode': {'1:0': {'type': 'VARIABLE_ALIAS', 'id': 'VariableID:1:1'}}
    }}}}
    m = normalize_variables_payload(raw)
    assert m['VariableID:5:5']['value'] == '→ VariableID:1:1'
