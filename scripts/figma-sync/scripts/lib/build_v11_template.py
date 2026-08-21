"""v11 (figma-sync 専用 6 sheet) テンプレを programmatic に生成する。

v5 テンプレの header / section 構造だけ模倣し、データ行は空にしておく。
excel_fill.py が実データで埋める。

Usage:
    python build_v11_template.py <output_path>
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
SECTION_FILL = PatternFill(start_color='FFE7F1FA', end_color='FFE7F1FA', fill_type='solid')
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=16)
SECTION_FONT = Font(bold=True, size=12)
THIN = Side(border_style='thin', color='FFBFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER = Alignment(horizontal='center', vertical='center')


def _hdr(ws, row, values, fill=HEADER_FILL):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i)
        c.value = v
        c.font = BOLD
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER


def _widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_01_cover(wb):
    ws = wb.create_sheet('01_表紙サマリ')
    ws['A1'] = 'figma-sync 実施報告書'
    ws['A1'].font = TITLE
    ws['A2'] = 'react_sudoku_app — Figma→React 同期パイプライン E2E 実施報告'
    ws['A2'].font = Font(italic=True)

    # プロジェクト情報 (行 4-15)
    info_rows = [
        ('プロジェクト名', ''),
        ('Figma ファイル ID', ''),
        ('Figma ファイル URL', ''),
        ('実施日 / run ts', ''),
        ('実施者', ''),
        ('final status', ''),
        ('検出方式', ''),
        ('コミット', ''),
        ('変更ファイル数', ''),
        ('成果 PR', ''),
        ('SKILL.md version', ''),
        ('備考', ''),
    ]
    for i, (label, val) in enumerate(info_rows, start=4):
        ws.cell(i, 1).value = label
        ws.cell(i, 1).font = BOLD
        ws.cell(i, 4).value = val

    # 【テスト実施サマリ】(行 17-)
    ws.cell(17, 1).value = '【テスト実施サマリ】'
    ws.cell(17, 1).font = SECTION_FONT
    ws.cell(17, 1).fill = SECTION_FILL
    _hdr(ws, 18, ['#', 'テストタイプ', '件数', 'ツール / 方法', '', '', '結果', ''])
    # データ行は excel_fill.py で埋める（行 19-25 を用意）

    ws.cell(26, 1).value = '合計'
    ws.cell(26, 1).font = BOLD

    # 【Figma 検出変更 概要】
    ws.cell(28, 1).value = '【Figma 検出変更 概要】'
    ws.cell(28, 1).font = SECTION_FONT
    ws.cell(28, 1).fill = SECTION_FILL
    _hdr(ws, 29, ['#', 'Component', 'Node ID', 'File', '変更 kind', '結果', '', ''])

    ws.cell(36, 1).value = 'KEY DISCOVERY:'
    ws.cell(36, 1).font = BOLD

    _widths(ws, {'A': 22, 'B': 22, 'C': 18, 'D': 40, 'E': 18, 'F': 12, 'G': 20, 'H': 20})
    return ws


def build_sheet_02_visual(wb):
    ws = wb.create_sheet('02_視覚エビデンス')
    ws['A1'] = '視覚エビデンス — Figma スクリーンショット'
    ws['A1'].font = TITLE
    ws['A2'] = 'BEFORE = apply 前、AFTER = apply 後（Playwright pixelmatch）'
    ws.cell(4, 2).value = 'BEFORE'
    ws.cell(4, 4).value = 'AFTER'
    ws.cell(4, 2).font = BOLD
    ws.cell(4, 4).font = BOLD
    # 行 6+ = 各 design change ブロック。excel_fill.py が動的に追加
    _widths(ws, {'A': 22, 'B': 42, 'C': 6, 'D': 42, 'E': 20})
    return ws


def build_sheet_03_state_diff(wb):
    ws = wb.create_sheet('03_state.json差分')
    ws['A1'] = 'state.json 差分 — apply 前/後 比較'
    ws['A1'].font = TITLE
    ws['A2'] = '.figma-sync-state.json / snapshots/last-full.json 差分'

    ws.cell(4, 1).value = '◆ バージョン・タイムスタンプ'
    ws.cell(4, 1).font = SECTION_FONT
    ws.cell(4, 1).fill = SECTION_FILL
    _hdr(ws, 5, ['#', 'フィールド', '前回値 (Before)', '今回値 (After)', '変化', '説明'])

    ws.cell(9, 1).value = '◆ perFrameHash (フレーム別ハッシュ)'
    ws.cell(9, 1).font = SECTION_FONT
    ws.cell(9, 1).fill = SECTION_FILL
    _hdr(ws, 10, ['#', 'フレーム ID', '前回値 (Before)', '今回値 (After)', '変化', '判定'])

    ws.cell(15, 1).value = '◆ メタ情報 / changedSinceLastRun'
    ws.cell(15, 1).font = SECTION_FONT
    ws.cell(15, 1).fill = SECTION_FILL
    _hdr(ws, 16, ['#', 'フィールド', '前回値 (Before)', '今回値 (After)', '変化', '判定'])

    ws.cell(26, 1).value = '◆ snapshot 変化'
    ws.cell(26, 1).font = SECTION_FONT
    ws.cell(26, 1).fill = SECTION_FILL
    _hdr(ws, 27, ['#', 'フィールド', '前回値', '今回値', '変化', '判定'])

    _widths(ws, {'A': 5, 'B': 26, 'C': 32, 'D': 32, 'E': 10, 'F': 42})
    return ws


def build_sheet_04_code_changes(wb):
    ws = wb.create_sheet('04_コード変更')
    ws['A1'] = 'コード変更 — Figma 変更を起点とした差分'
    ws['A1'].font = TITLE

    ws.cell(3, 1).value = '◆ コミット / 適用ファイル一覧'
    ws.cell(3, 1).font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    _hdr(ws, 4, ['#', 'コミット', 'ファイル', '変更種別', '変更内容', 'Figma 起源', '判定'])

    ws.cell(9, 1).value = '◆ reconcile drift / warnings'
    ws.cell(9, 1).font = SECTION_FONT
    ws.cell(9, 1).fill = SECTION_FILL
    _hdr(ws, 10, ['#', '対象', '内容', '判断根拠', 'Figma 起源', '推奨対応', '結果'])

    ws.cell(15, 1).value = '合計'
    ws.cell(15, 1).font = BOLD
    _widths(ws, {'A': 5, 'B': 14, 'C': 40, 'D': 14, 'E': 44, 'F': 32, 'G': 12})
    return ws


def build_sheet_05_test_cases(wb):
    ws = wb.create_sheet('05_テストケース')
    ws['A1'] = 'テストケース — Phase 別実施ログ'
    ws['A1'].font = TITLE
    _hdr(ws, 2, [
        '', 'No.', 'システム / 機能名', 'テストケース\n(観点・前提条件・入力)',
        '期待結果', '実際結果', '開発時テスト\n(Claude)', '実施者',
        '追加テスト (受入相当)\n(Bridge)', '実施者', '実行ログ (エビデンス)'
    ])
    ws.cell(3, 7).value = '実施日'
    ws.cell(3, 8).value = '実施者'
    ws.cell(3, 9).value = '実施日'
    ws.cell(3, 10).value = '実施者'
    for col in range(7, 11):
        ws.cell(3, col).font = BOLD
        ws.cell(3, col).fill = HEADER_FILL

    # excel_fill.py が動的に Phase 別 TC 行を挿入（section header + rows）
    _widths(ws, {'A': 3, 'B': 8, 'C': 20, 'D': 40, 'E': 30, 'F': 40,
                 'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 30})
    ws.row_dimensions[2].height = 42
    return ws


def build_sheet_06_summary(wb):
    ws = wb.create_sheet('06_総合結果')
    ws['A1'] = '総合結果サマリ'
    ws['A1'].font = TITLE

    ws.cell(3, 1).value = '◆ PR / commit 確認'
    ws.cell(3, 1).font = SECTION_FONT
    ws.cell(3, 1).fill = SECTION_FILL
    _hdr(ws, 4, ['項目', '内容', '詳細', '結果'])

    # 行 5-11 = PR 確認 rows (excel_fill.py が埋める)

    ws.cell(12, 1).value = '◆ 最終判定'
    ws.cell(12, 1).font = SECTION_FONT
    ws.cell(12, 1).fill = SECTION_FILL
    _hdr(ws, 13, ['検証項目', '結果サマリ', '根拠', '判定'])

    # 行 14-24 = 判定 rows

    ws.cell(25, 1).value = '◆ 次回アクション / 申し送り'
    ws.cell(25, 1).font = SECTION_FONT
    ws.cell(25, 1).fill = SECTION_FILL
    _hdr(ws, 26, ['#', '項目', '詳細', '担当 / 期限'])

    ws.cell(35, 1).value = '総合判定:'
    ws.cell(35, 1).font = BOLD
    ws.cell(37, 1).value = '作成:'
    ws.cell(37, 1).font = BOLD
    _widths(ws, {'A': 32, 'B': 32, 'C': 60, 'D': 14})
    return ws


def build(output_path: str):
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)
    build_sheet_01_cover(wb)
    build_sheet_02_visual(wb)
    build_sheet_03_state_diff(wb)
    build_sheet_04_code_changes(wb)
    build_sheet_05_test_cases(wb)
    build_sheet_06_summary(wb)
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        out = Path(__file__).resolve().parents[2] / 'templates' / 'sudoku_figma-sync実施報告書_2026-08-21_v11.xlsx'
    else:
        out = Path(sys.argv[1])
    out.parent.mkdir(parents=True, exist_ok=True)
    build(str(out))
    print(f'Generated: {out}')
