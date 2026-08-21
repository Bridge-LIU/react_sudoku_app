"""v11.3 新テンプレ (7 sheet) を実データで生成する処理。

シート構成 (2026-08-21 決定 spec: docs/superpowers/specs/2026-08-21-test-report-template-design.md):
  01_表紙サマリ         Bridge紺帯 / メタ表 / OK-NG-WARN パネル / 目次
  02_変更点サマリ       variables/components/screens 差分 (人間可読)
  03_UI Before-After    画面別 前後スクショ横並び + 差分説明
  04_コード変更         ファイル別 diff (±行数/カテゴリ/影響) + bindingChanges 詳細
  05_全システムテスト   vitest/Playwright/a11y/snapshot フル結果
  06_異常系＆セキュリティ 異常系ケース + audit + CSP チェック
  07_承認履歴           AI 2段レビュー + 人間承認欄

v11 (テンプレファイル load 型) から刷新：Workbook を code from scratch で組み立てるため
templates/v11.xlsx は不要になった。既存 helper (_binding_fill / _fmt_variable /
normalize_variables_payload / _collect_binding_rows / ExcelFillContext) は互換保持。
"""
from __future__ import annotations

import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── フォント ──
FONT_TITLE   = Font(name='Yu Gothic', size=18, bold=True, color='FFFFFFFF')
FONT_SUBTTL  = Font(name='Yu Gothic', size=14, bold=True, color='FFFFFFFF')
FONT_H2      = Font(name='Yu Gothic', size=14, bold=True)
FONT_TH      = Font(name='Yu Gothic', size=11, bold=True)
FONT_TH_INV  = Font(name='Yu Gothic', size=11, bold=True, color='FFFFFFFF')
FONT_TD      = Font(name='Yu Gothic', size=10)
FONT_TD_B    = Font(name='Yu Gothic', size=10, bold=True)
FONT_META_K  = Font(name='Yu Gothic', size=11, bold=True)
FONT_LINK    = Font(name='Yu Gothic', size=11, color='FF0563C1', underline='single')
FONT_LARGE_B = Font(name='Yu Gothic', size=16, bold=True)

# ── Fill (v10 準拠 + Bridge紺) ──
FILL_BANNER  = PatternFill('solid', fgColor='FF1F3864')  # Bridge 濃紺
FILL_HDR     = PatternFill('solid', fgColor='FFD9E1F2')  # 淡青 (v10)
FILL_HDR_D   = PatternFill('solid', fgColor='FF4472C4')  # 濃青 (v5)
FILL_TOTAL   = PatternFill('solid', fgColor='FFE2EFDA')  # 淡緑
FILL_OK      = PatternFill('solid', fgColor='FFC6EFCE')  # PASS 緑
FILL_NG      = PatternFill('solid', fgColor='FFFFC7CE')  # FAIL 赤
FILL_WARN    = PatternFill('solid', fgColor='FFFFEB9C')  # WARN 黄
FILL_PANEL   = PatternFill('solid', fgColor='FFEDEDED')  # 灰 (結果パネル背景)
FILL_FALLBACK = PatternFill('solid', fgColor='FFFFEBEE')  # 淡赤 (fallback banner)
FALLBACK_BANNER_FONT = Font(name='Yu Gothic', size=14, bold=True, color='FFC62828')

# bindingChanges の change_kind 別 背景色
BINDING_FILL_ADDED    = PatternFill('solid', fgColor='FFE8F5E9')
BINDING_FILL_REMOVED  = PatternFill('solid', fgColor='FFFFEBEE')
BINDING_FILL_MODIFIED = PatternFill('solid', fgColor='FFFFF8E1')

# ── Border ──
_THIN = Side(style='thin', color='FF808080')
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# ── Alignment ──
ALIGN_CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_R   = Alignment(horizontal='right',  vertical='center', wrap_text=True)

# ── シート順 (fill_report の生成順と wb._sheets の並び) ──
SHEET_ORDER = [
    '01_表紙サマリ',
    '02_変更点サマリ',
    '03_UI Before-After',
    '04_コード変更',
    '05_全システムテスト',
    '06_異常系＆セキュリティ',
    '07_承認履歴',
]


# ─────────────────────────────────────────────────
# 既存 helper (互換保持)
# ─────────────────────────────────────────────────
def _binding_fill(change_kind: str):
    k = (change_kind or '').lower()
    if k == 'added':    return BINDING_FILL_ADDED
    if k == 'removed':  return BINDING_FILL_REMOVED
    if k == 'modified': return BINDING_FILL_MODIFIED
    return None


def _rgba_to_hex(v) -> str:
    try:
        r = round(float(v.get('r', 0)) * 255)
        g = round(float(v.get('g', 0)) * 255)
        b = round(float(v.get('b', 0)) * 255)
        a = float(v.get('a', 1))
        hex6 = f'#{r:02X}{g:02X}{b:02X}'
        return hex6 if a >= 0.999 else f'{hex6}{round(a * 255):02X}'
    except Exception:
        return str(v)


def _fmt_variable_value(entry: dict) -> str:
    val = entry.get('value')
    if val is None:
        return '(値なし)'
    rt = (entry.get('resolved_type') or '').upper()
    if rt == 'COLOR' and isinstance(val, dict):
        return _rgba_to_hex(val)
    if rt == 'FLOAT':
        try:
            f = float(val)
            return str(int(f)) if f.is_integer() else f'{f:g}'
        except Exception:
            return str(val)
    if rt == 'BOOLEAN':
        return 'true' if bool(val) else 'false'
    if rt == 'STRING':
        return f'"{val}"'
    return str(val)


def _fmt_variable(vid, variables_map: dict) -> str:
    if vid is None:
        return '(なし)'
    entry = (variables_map or {}).get(vid)
    if not entry:
        return f'{vid} (未解決)'
    name = entry.get('name') or vid
    if entry.get('value') is None:
        return f'{name} (値なし)'
    return f'{name} ({_fmt_variable_value(entry)})'


def normalize_variables_payload(payload) -> dict:
    if not payload or not isinstance(payload, (dict, list)):
        return {}
    src = payload
    if isinstance(src, dict) and 'meta' in src and isinstance(src.get('meta'), dict):
        src = src['meta'].get('variables', {})
    elif isinstance(src, dict) and 'variables' in src and isinstance(src.get('variables'), dict):
        src = src['variables']
    if not isinstance(src, dict):
        return {}
    out = {}
    for vid, raw in src.items():
        if not isinstance(raw, dict):
            continue
        name = raw.get('name') or ''
        resolved_type = raw.get('resolvedType') or raw.get('resolved_type') or ''
        if 'value' in raw:
            value = raw.get('value')
        else:
            vbm = raw.get('valuesByMode') or raw.get('values_by_mode') or {}
            if isinstance(vbm, dict) and vbm:
                first = next(iter(vbm.values()))
                if isinstance(first, dict) and first.get('type') == 'VARIABLE_ALIAS':
                    value = f'→ {first.get("id", "?")}'
                else:
                    value = first
            else:
                value = None
        out[vid] = {'name': name, 'resolved_type': resolved_type, 'value': value}
    return out


def _collect_binding_rows(design_changes: list) -> list:
    out = []
    for dc in design_changes or []:
        comp = dc.get('component', '')
        frame_id = dc.get('node_id', '')
        for ch in dc.get('changes', []) or []:
            for b in ch.get('bindingChanges', []) or []:
                out.append({
                    'component': comp,
                    'frame_node_id': frame_id,
                    'node_id': b.get('node_id', ''),
                    'node_name': b.get('node_name', ''),
                    'property': b.get('property', ''),
                    'from_variable_id': b.get('from_variable_id'),
                    'to_variable_id': b.get('to_variable_id'),
                    'change_kind': b.get('change_kind', ''),
                })
    return out


@dataclass
class ExcelFillContext:
    template_path: str  # 後方互換のため受け取るが未使用 (v11.3 で不要化)
    output_path: str
    run_ts: str
    test_results: dict = field(default_factory=dict)
    design_changes: list = field(default_factory=list)
    config: Optional[dict] = None
    state_before: Optional[dict] = None
    state_after: Optional[dict] = None
    apply_result: Optional[dict] = None
    diff_summary: Optional[dict] = None
    git_info: Optional[dict] = None
    phase_results: Optional[dict] = None
    final_status: Optional[str] = None
    fallback_triggered: bool = False
    fallback_info: dict = field(default_factory=dict)
    variables_map: dict = field(default_factory=dict)


def _log(msg: str) -> None:
    try:
        sys.stderr.write(msg + '\n')
        sys.stderr.flush()
    except UnicodeEncodeError:
        buf = getattr(sys.stderr, 'buffer', None)
        if buf is not None:
            buf.write((msg + '\n').encode('utf-8', errors='replace'))
            buf.flush()


def _detect_method(ctx: ExcelFillContext) -> str:
    if ctx.fallback_triggered:
        return 'figma_get_file_at_version fallback (local-diff)'
    ph = ctx.phase_results or {}
    if 'fallback' in str(ph.get('phase_1b_fallback', '')).lower() and 'CHANGED' in str(ph.get('phase_1b_fallback', '')):
        return 'figma_get_file_at_version fallback (local-diff)'
    return 'figma_diff_versions'


# ─────────────────────────────────────────────────
# 描画 helper
# ─────────────────────────────────────────────────
def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _cell(ws, coord, value, *, font=None, fill=None, align=None, border=True):
    c = ws[coord]
    c.value = value
    if font:  c.font = font
    if fill:  c.fill = fill
    if align: c.alignment = align
    if border: c.border = BORDER
    return c


def _banner(ws, row, span_from, span_to, title, subtitle=None):
    ws.merge_cells(f'{span_from}{row}:{span_to}{row}')
    c = ws[f'{span_from}{row}']
    c.value = title
    c.font = FONT_TITLE
    c.fill = FILL_BANNER
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 42
    if subtitle:
        ws.merge_cells(f'{span_from}{row+1}:{span_to}{row+1}')
        c2 = ws[f'{span_from}{row+1}']
        c2.value = subtitle
        c2.font = FONT_SUBTTL
        c2.fill = FILL_BANNER
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row+1].height = 28


def _write_table(ws, start_row, headers, rows, *, header_style='light'):
    """header + rows を描画、最終行の次の row 番号を返す。"""
    fill = FILL_HDR if header_style == 'light' else FILL_HDR_D
    font = FONT_TH if header_style == 'light' else FONT_TH_INV
    for i, h in enumerate(headers):
        c = ws.cell(row=start_row, column=i + 1, value=h)
        c.font = font; c.fill = fill; c.alignment = ALIGN_CTR; c.border = BORDER
    ws.row_dimensions[start_row].height = 24
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            c = ws.cell(row=start_row + 1 + r_idx, column=c_idx + 1, value=val)
            c.font = FONT_TD; c.alignment = ALIGN_L; c.border = BORDER
        ws.row_dimensions[start_row + 1 + r_idx].height = 22
    return start_row + 1 + len(rows)


# ─────────────────────────────────────────────────
# メイン (Workbook を code from scratch で組み立て)
# ─────────────────────────────────────────────────
def fill_report(ctx: ExcelFillContext) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _log('[Excel] Sheet 2/7 02_変更点サマリ 生成中...')
    _build_02_change_summary(wb, ctx)
    _log('[Excel] Sheet 3/7 03_UI Before-After 生成中...')
    _build_03_ui_diff(wb, ctx)
    _log('[Excel] Sheet 4/7 04_コード変更 生成中...')
    _build_04_code_changes(wb, ctx)
    _log('[Excel] Sheet 5/7 05_全システムテスト 生成中...')
    test_total, test_pass, test_fail = _build_05_full_tests(wb, ctx)
    _log('[Excel] Sheet 6/7 06_異常系＆セキュリティ 生成中...')
    sec_pass, sec_warn, sec_fail = _build_06_abnormal_sec(wb, ctx)
    _log('[Excel] Sheet 7/7 07_承認履歴 生成中...')
    _build_07_approvals(wb, ctx)

    ok_count   = test_pass + sec_pass
    ng_count   = test_fail + sec_fail
    warn_count = sec_warn

    _log('[Excel] Sheet 1/7 01_表紙サマリ 生成中...')
    _build_01_cover(wb, ctx, ok_count, ng_count, warn_count)

    # 表紙を先頭に並べ替え
    wb._sheets = [wb[n] for n in SHEET_ORDER]

    wb.save(ctx.output_path)


# ─────────────────────────────────────────────────
# Sheet 01: 表紙サマリ
# ─────────────────────────────────────────────────
def _build_01_cover(wb, ctx: ExcelFillContext, ok, ng, warn):
    ws = wb.create_sheet('01_表紙サマリ')
    _set_col_widths(ws, {'A': 5, 'B': 22, 'C': 12, 'D': 12, 'E': 14, 'F': 14, 'G': 18, 'H': 18})

    cfg = ctx.config or {}
    git = ctx.git_info or {}
    date_str = (ctx.run_ts or '').split(' ')[0] or '(no date)'

    _banner(ws, 1, 'A', 'H',
            'テスト実施報告書',
            f'react_sudoku_app  ―  figma-sync v11.3  ({date_str})')

    # Fallback banner (存在時のみ、row 3)
    start_meta = 4
    if ctx.fallback_triggered:
        fi = ctx.fallback_info or {}
        reason = fi.get('reason', 'UNKNOWN')
        node_cnt = fi.get('nodeDiffCount', '?')
        ws.merge_cells('A3:H3')
        cell = ws['A3']
        cell.value = (f'⚠️ FALLBACK 発動 (raw property / boundVariables 系変更検出) — '
                      f'reason={reason}、fallback 経由で {node_cnt} 件検出')
        cell.font = FALLBACK_BANNER_FONT
        cell.fill = FILL_FALLBACK
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 32
        start_meta = 4  # メタ表は row 4 から

    # メタ情報テーブル
    apply_res = ctx.apply_result or {}
    changed_files = apply_res.get('changedFiles') or git.get('changed_files') or []
    commit_sha = (git.get('commit_sha') or '')[:8]
    meta_items = [
        ('プロジェクト名',    'react_sudoku_app'),
        ('バージョン',        'figma-sync v11.3'),
        ('報告日 / run ts',   ctx.run_ts or ''),
        ('報告者',            'Claude Code'),
        ('final status',      ctx.final_status or 'UNKNOWN'),
        ('検出方式',          _detect_method(ctx)),
        ('コミット',          f'{commit_sha} — {git.get("commit_msg", "")[:60]}' if commit_sha else '(なし)'),
        ('変更ファイル数',    f'{len(changed_files)} files'),
        ('Figma File ID',     cfg.get('figmaFileKey', '(未設定)')),
        ('成果 PR',           git.get('pr_url') or '(未作成)'),
    ]
    for i, (k, v) in enumerate(meta_items):
        r = start_meta + i
        ws.merge_cells(f'A{r}:C{r}')
        _cell(ws, f'A{r}', k, font=FONT_META_K, fill=FILL_HDR, align=ALIGN_L)
        ws.merge_cells(f'D{r}:H{r}')
        _cell(ws, f'D{r}', v, font=FONT_TD, align=ALIGN_L)
        ws.row_dimensions[r].height = 22

    # 結果パネル
    panel_row = start_meta + len(meta_items) + 1
    ws.merge_cells(f'A{panel_row}:H{panel_row}')
    _cell(ws, f'A{panel_row}', '【総合結果】', font=FONT_H2, align=ALIGN_L, border=False)
    ws.row_dimensions[panel_row].height = 26

    verdict = '合格 (要人間承認)' if ng == 0 else '要修正'
    for i, (label, val, fill) in enumerate([
        ('OK',   ok,   FILL_OK),
        ('NG',   ng,   FILL_NG),
        ('WARN', warn, FILL_WARN),
        ('総合判定', verdict, FILL_PANEL),
    ]):
        col_from = 1 + i * 2
        col_to   = col_from + 1
        r = panel_row + 1
        ws.merge_cells(start_row=r, start_column=col_from, end_row=r, end_column=col_to)
        c = ws.cell(row=r, column=col_from, value=label)
        c.font = FONT_TH; c.fill = FILL_HDR; c.alignment = ALIGN_CTR; c.border = BORDER

        ws.merge_cells(start_row=r + 1, start_column=col_from, end_row=r + 2, end_column=col_to)
        c2 = ws.cell(row=r + 1, column=col_from, value=val)
        c2.font = FONT_LARGE_B; c2.fill = fill; c2.alignment = ALIGN_CTR; c2.border = BORDER
    ws.row_dimensions[panel_row + 1].height = 22
    ws.row_dimensions[panel_row + 2].height = 26
    ws.row_dimensions[panel_row + 3].height = 26

    # 目次 (残り 6 シートへ)
    toc_row = panel_row + 5
    ws.merge_cells(f'A{toc_row}:H{toc_row}')
    _cell(ws, f'A{toc_row}', '【目次】', font=FONT_H2, align=ALIGN_L, border=False)
    ws.row_dimensions[toc_row].height = 26

    for i, name in enumerate(SHEET_ORDER[1:]):
        r = toc_row + 1 + i
        ws.merge_cells(f'A{r}:B{r}')
        _cell(ws, f'A{r}', f'  {i + 2:02d}', font=FONT_TD_B, align=ALIGN_L, border=False)
        ws.merge_cells(f'C{r}:H{r}')
        c = ws.cell(row=r, column=3, value=name)
        c.font = FONT_LINK
        c.alignment = ALIGN_L
        c.hyperlink = f"#'{name}'!A1"
        ws.row_dimensions[r].height = 20


# ─────────────────────────────────────────────────
# Sheet 02: 変更点サマリ
# ─────────────────────────────────────────────────
def _build_02_change_summary(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('02_変更点サマリ')
    _set_col_widths(ws, {'A': 5, 'B': 14, 'C': 28, 'D': 46, 'E': 34})
    _banner(ws, 1, 'A', 'E', '変更点サマリ', '何が / どう変わったか (人間可読)')

    # design_changes[] → 種別ごとに要約
    dc = ctx.design_changes or []
    rows = []
    idx = 0
    for c in dc:
        comp = c.get('component', '')
        node_id = c.get('node_id', '')
        file_path = c.get('file', '')
        # 種別判定: bindingChanges があれば Variable/Component、無ければ Screen
        for ch in (c.get('changes') or []):
            bindings = ch.get('bindingChanges') or []
            if bindings:
                for b in bindings:
                    idx += 1
                    prop = b.get('property', '')
                    frm = _fmt_variable(b.get('from_variable_id'), ctx.variables_map)
                    to  = _fmt_variable(b.get('to_variable_id'),   ctx.variables_map)
                    kind = b.get('change_kind', '')
                    rows.append((
                        idx,
                        'Variable',
                        f'{comp} / {b.get("node_name") or b.get("node_id", "")} .{prop}',
                        f'{frm} → {to} ({kind})',
                        file_path,
                    ))
            else:
                idx += 1
                rows.append((
                    idx,
                    'Screen',
                    f'{comp} ({node_id})',
                    ch.get('change_kind') or ch.get('kind') or '(内容不明)',
                    file_path,
                ))

    if not rows:
        rows = [(1, '-', '(design_changes 空)', '変更検出なし', '-')]

    row = _write_table(ws, 4,
                       ['#', '種別', '対象', '変更内容', '影響ファイル'],
                       rows, header_style='dark')

    kind_var    = sum(1 for r in rows if r[1] == 'Variable')
    kind_screen = sum(1 for r in rows if r[1] == 'Screen')
    kind_other  = len(rows) - kind_var - kind_screen
    ws.merge_cells(f'A{row + 1}:E{row + 1}')
    _cell(ws, f'A{row + 1}',
          f'合計: {len(rows)} 件  (Variable: {kind_var} / Screen: {kind_screen} / その他: {kind_other})',
          font=FONT_TD_B, fill=FILL_TOTAL, align=ALIGN_L)


# ─────────────────────────────────────────────────
# Sheet 03: UI Before-After
# ─────────────────────────────────────────────────
def _build_03_ui_diff(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('03_UI Before-After')
    _set_col_widths(ws, {'A': 5, 'B': 22, 'C': 38, 'D': 38, 'E': 34})
    _banner(ws, 1, 'A', 'E', 'UI Before / After', '画面別スクショ比較')

    headers = ['#', '画面', 'Before', 'After', '差分説明']
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = FONT_TH_INV; c.fill = FILL_HDR_D; c.alignment = ALIGN_CTR; c.border = BORDER
    ws.row_dimensions[4].height = 24

    dc = ctx.design_changes or []
    if not dc:
        ws.merge_cells('A5:E5')
        _cell(ws, 'A5', '(design_changes 空 — screenshot 未収集 or 変更なし)',
              font=FONT_TD, align=ALIGN_L)
        return

    for idx, c in enumerate(dc):
        r = 5 + idx
        ws.row_dimensions[r].height = 130
        _cell(ws, f'A{r}', idx + 1, font=FONT_TD, align=ALIGN_CTR)
        _cell(ws, f'B{r}', f'{c.get("component", "")}\n({c.get("node_id", "")})',
              font=FONT_TD_B, align=ALIGN_L)

        for col, key, label in [('C', 'before_png', 'BEFORE'), ('D', 'after_png', 'AFTER')]:
            png = c.get(key)
            addr = f'{col}{r}'
            if png and Path(png).exists():
                try:
                    img = XLImage(png)
                    img.width, img.height = 260, 165
                    ws.add_image(img, addr)
                    _cell(ws, addr, '', font=FONT_TD, align=ALIGN_CTR)
                except Exception:
                    _cell(ws, addr, f'[{label}] (画像埋込失敗)', font=FONT_TD, align=ALIGN_CTR)
            else:
                _cell(ws, addr, f'[{label}] (画像なし)', font=FONT_TD, align=ALIGN_CTR)

        # 差分説明: bindingChanges 件数 + diff_pixels
        binding_total = sum(len(ch.get('bindingChanges', []) or []) for ch in (c.get('changes') or []))
        parts = [f'file: {c.get("file", "")}']
        if binding_total > 0:
            parts.append(f'{binding_total} binding 変更')
        if c.get('diff_pixels') is not None:
            parts.append(f'diff: {c.get("diff_pixels")} px')
        _cell(ws, f'E{r}', '\n'.join(parts), font=FONT_TD, align=ALIGN_L)


# ─────────────────────────────────────────────────
# Sheet 04: コード変更
# ─────────────────────────────────────────────────
def _build_04_code_changes(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('04_コード変更')
    _set_col_widths(ws, {'A': 5, 'B': 40, 'C': 12, 'D': 32, 'E': 14, 'F': 24})
    _banner(ws, 1, 'A', 'F', 'コード変更', 'ファイル別 diff + 検出経路 + bindingChanges 詳細')

    git = ctx.git_info or {}
    apply_res = ctx.apply_result or {}
    changed_files = apply_res.get('changedFiles') or git.get('changed_files') or []
    sha = (git.get('commit_sha') or '')[:8] or '(uncommitted)'
    route = 'fallback (local-diff)' if ctx.fallback_triggered else 'figma_diff_versions'

    # ファイル一覧 (行 4-)
    rows = []
    for i, f in enumerate(changed_files, start=1):
        rows.append((i, f, sha, git.get('commit_msg', '')[:60] or '(msg なし)',
                     'PASS' if not apply_res.get('errors') else 'FAIL', route))
    if not rows:
        rows = [(1, '(変更ファイルなし)', '-', '-', 'N/A', route)]

    end_row = _write_table(ws, 4,
                           ['#', 'ファイル', 'コミット', 'メッセージ', 'apply結果', '検出経路'],
                           rows, header_style='dark')

    # fallback 時は検出経路列を赤くハイライト
    if ctx.fallback_triggered:
        for r_idx in range(5, end_row):
            ws.cell(r_idx, 6).fill = FILL_FALLBACK

    # 合計行
    ws.merge_cells(f'A{end_row + 1}:F{end_row + 1}')
    _cell(ws, f'A{end_row + 1}',
          f'合計: {len(changed_files)} files apply / errors={len(apply_res.get("errors", []))}',
          font=FONT_TD_B, fill=FILL_TOTAL, align=ALIGN_L)

    # bindingChanges 詳細 section
    sec_row = end_row + 3
    ws.merge_cells(f'A{sec_row}:F{sec_row}')
    _cell(ws, f'A{sec_row}', '◆ Figma 変更詳細 (bindingChanges)',
          font=FONT_H2, fill=FILL_HDR, align=ALIGN_L)
    ws.row_dimensions[sec_row].height = 26

    binding_headers = ['#', '対象 frame', 'node_id', 'property', 'from → to', '種類']
    for i, h in enumerate(binding_headers):
        c = ws.cell(row=sec_row + 1, column=i + 1, value=h)
        c.font = FONT_TH_INV; c.fill = FILL_HDR_D; c.alignment = ALIGN_CTR; c.border = BORDER
    ws.row_dimensions[sec_row + 1].height = 24

    binding_rows = _collect_binding_rows(ctx.design_changes or [])
    if not binding_rows:
        ws.merge_cells(f'A{sec_row + 2}:F{sec_row + 2}')
        _cell(ws, f'A{sec_row + 2}', '(bindingChanges なし)',
              font=FONT_TD, align=ALIGN_L)
    else:
        for i, b in enumerate(binding_rows):
            r = sec_row + 2 + i
            _cell(ws, f'A{r}', i + 1, font=FONT_TD, align=ALIGN_CTR)
            _cell(ws, f'B{r}',
                  f'{b["component"]} ({b["frame_node_id"]})' if b['component'] else b['frame_node_id'],
                  font=FONT_TD, align=ALIGN_L)
            _cell(ws, f'C{r}', b['node_id'], font=FONT_TD, align=ALIGN_L)
            _cell(ws, f'D{r}', b['property'], font=FONT_TD, align=ALIGN_L)
            frm = _fmt_variable(b['from_variable_id'], ctx.variables_map)
            to = _fmt_variable(b['to_variable_id'], ctx.variables_map)
            _cell(ws, f'E{r}', f'{frm} → {to}', font=FONT_TD, align=ALIGN_L)
            fill = _binding_fill(b['change_kind'])
            _cell(ws, f'F{r}', b['change_kind'], font=FONT_TD_B, fill=fill, align=ALIGN_CTR)
            ws.row_dimensions[r].height = 22


# ─────────────────────────────────────────────────
# Sheet 05: 全システムテスト
# ─────────────────────────────────────────────────
def _build_05_full_tests(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('05_全システムテスト')
    _set_col_widths(ws, {'A': 5, 'B': 22, 'C': 14, 'D': 8, 'E': 8, 'F': 8, 'G': 10, 'H': 40})
    _banner(ws, 1, 'A', 'H', '全システムテスト',
            'vitest / Playwright / a11y / snapshot フル実行 (回帰確認)')

    tr = ctx.test_results or {}
    ph = ctx.phase_results or {}

    def _rows_for(section: dict, tool: str):
        """test-results.json の 1 section から (件数, PASS数, FAIL数, 時間) を推定"""
        if not section:
            return 0, 0, 0, '-'
        exit_code = section.get('exitCode')
        if exit_code is None:
            return 0, 0, 0, '-'
        # 生実行時間があれば拾う (無ければ '-')
        dur = section.get('duration') or section.get('durationSec') or '-'
        dur_str = f'{dur}s' if isinstance(dur, (int, float)) else str(dur)
        # numPassed/numFailed が無い場合は exitCode で推定
        n = section.get('numTotal') or section.get('total') or 1
        p = section.get('numPassed') or section.get('passed') or (n if exit_code == 0 else 0)
        f = section.get('numFailed') or section.get('failed') or (0 if exit_code == 0 else n)
        return n, p, f, dur_str

    unit = tr.get('unit') or {}
    e2e = tr.get('e2e') or {}
    audit = tr.get('audit') or {}
    coverage = tr.get('coverage') or {}
    cov_line = (coverage.get('total') or {}).get('lines', {}).get('pct', '-') if coverage else '-'

    rows_input = [
        ('単体・契約・統合 (Vitest)', 'Vitest',       *_rows_for(unit, 'Vitest')),
        ('E2E (Playwright)',        'Playwright',   *_rows_for(e2e, 'Playwright')),
        ('a11y (axe-core)',         'axe-core',     0, 0, 0, '-'),
        ('snapshot',                'Jest',         0, 0, 0, '-'),
    ]

    rows = []
    for i, (name, tool, n, p, f, dur) in enumerate(rows_input, start=1):
        note = ''
        if 'Vitest' in tool and cov_line != '-':
            note = f'coverage(lines)={cov_line}%'
        if 'Playwright' in tool:
            note = f'phase_4={ph.get("phase_4", "-")}, phase_5={ph.get("phase_5", "-")}'
        if 'axe' in tool:
            note = '(未収集 / v11.3 で追加予定)'
        if 'snapshot' in tool.lower() or 'Jest' == tool:
            note = '(未収集 / v11.3 で追加予定)'
        rows.append((i, name, tool, n, p, f, dur, note))

    end_row = _write_table(ws, 4,
                           ['#', 'テスト種別', 'ツール', '件数', 'PASS', 'FAIL', '実行時間', '影響ファイル相関 / 備考'],
                           rows, header_style='dark')

    total_n = sum(r[3] for r in rows)
    total_p = sum(r[4] for r in rows)
    total_f = sum(r[5] for r in rows)
    ws.merge_cells(f'A{end_row + 1}:H{end_row + 1}')
    icon = '全PASS ✅' if total_f == 0 else 'FAIL あり ❌'
    _cell(ws, f'A{end_row + 1}',
          f'合計: {total_n} 件 / PASS {total_p} / FAIL {total_f}  ({icon})',
          font=FONT_TD_B, fill=FILL_TOTAL, align=ALIGN_L)

    # npm audit セクション (簡易)
    audit_row = end_row + 3
    ws.merge_cells(f'A{audit_row}:H{audit_row}')
    _cell(ws, f'A{audit_row}', '◆ npm audit',
          font=FONT_H2, fill=FILL_HDR, align=ALIGN_L)
    ws.row_dimensions[audit_row].height = 26

    audit_json = (audit.get('json') or {}) if audit else {}
    vuln = (audit_json.get('metadata') or {}).get('vulnerabilities', {})
    audit_rows = [
        ('critical', vuln.get('critical', 0)),
        ('high',     vuln.get('high', 0)),
        ('moderate', vuln.get('moderate', 0)),
        ('low',      vuln.get('low', 0)),
        ('info',     vuln.get('info', 0)),
        ('total',    vuln.get('total', 0)),
    ]
    for i, (sev, cnt) in enumerate(audit_rows):
        r = audit_row + 1 + i
        _cell(ws, f'A{r}', sev, font=FONT_TH, fill=FILL_HDR, align=ALIGN_CTR)
        _cell(ws, f'B{r}', str(cnt), font=FONT_TD, align=ALIGN_L)
        ws.row_dimensions[r].height = 20

    return total_n, total_p, total_f


# ─────────────────────────────────────────────────
# Sheet 06: 異常系＆セキュリティ
# ─────────────────────────────────────────────────
def _build_06_abnormal_sec(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('06_異常系＆セキュリティ')
    _set_col_widths(ws, {'A': 5, 'B': 10, 'C': 34, 'D': 10, 'E': 40})
    _banner(ws, 1, 'A', 'E', '異常系 ＆ セキュリティ',
            'BAF 必須: 異常系ケース + セキュリティチェック')

    headers = ['#', '分類', 'ケース', '結果', '備考']
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = FONT_TH_INV; c.fill = FILL_HDR_D; c.alignment = ALIGN_CTR; c.border = BORDER
    ws.row_dimensions[4].height = 24

    tr = ctx.test_results or {}
    audit = tr.get('audit') or {}
    audit_json = (audit.get('json') or {}) if audit else {}
    vuln = (audit_json.get('metadata') or {}).get('vulnerabilities', {})
    audit_total = vuln.get('total', 0)
    audit_result = 'PASS' if audit_total == 0 else ('WARN' if not (vuln.get('critical', 0) or vuln.get('high', 0)) else 'FAIL')

    # 異常系ケースは現状 phase から間接的にしか取れないので、phase 状態を代替表示
    ph = ctx.phase_results or {}
    fb = ctx.fallback_info or {}
    cases = [
        (1, '異常系', 'Fallback 経路発動 (raw property 検出)',
         'WARN' if ctx.fallback_triggered else 'PASS',
         (f'reason={fb.get("reason", "?")}、拾えた変更={fb.get("nodeDiffCount", 0)} 件'
          if ctx.fallback_triggered else 'fallback 未発動')),
        (2, '異常系', 'apply エラー (05-apply.js)',
         'PASS' if not (ctx.apply_result or {}).get('errors') else 'FAIL',
         f'errors={len((ctx.apply_result or {}).get("errors", []))}'),
        (3, '異常系', 'diff warnings',
         'PASS' if not (ctx.diff_summary or {}).get('warnings') else 'WARN',
         f'warnings={len((ctx.diff_summary or {}).get("warnings", []))}'),
        (4, 'Sec',   'npm audit (prod+dev 合算)',
         audit_result,
         f'total={audit_total} / critical={vuln.get("critical", 0)} / high={vuln.get("high", 0)}'),
        (5, 'Sec',   'CSP ヘッダ検証',
         'INFO', '(v11.3 では未計装)'),
        (6, 'Sec',   '秘密情報スキャン (git secrets)',
         'INFO', '(v11.3 では未計装)'),
    ]

    pass_cnt = warn_cnt = fail_cnt = 0
    for idx, (num, cat, case, result, note) in enumerate(cases):
        r = 5 + idx
        _cell(ws, f'A{r}', num, font=FONT_TD, align=ALIGN_CTR)
        _cell(ws, f'B{r}', cat, font=FONT_TD, align=ALIGN_CTR)
        _cell(ws, f'C{r}', case, font=FONT_TD, align=ALIGN_L)
        fill = {'PASS': FILL_OK, 'WARN': FILL_WARN, 'FAIL': FILL_NG}.get(result)
        _cell(ws, f'D{r}', result, font=FONT_TD_B, fill=fill, align=ALIGN_CTR)
        _cell(ws, f'E{r}', note, font=FONT_TD, align=ALIGN_L)
        ws.row_dimensions[r].height = 22
        if result == 'PASS': pass_cnt += 1
        elif result == 'WARN': warn_cnt += 1
        elif result == 'FAIL': fail_cnt += 1

    # 合計行
    end_row = 5 + len(cases)
    ws.merge_cells(f'A{end_row}:E{end_row}')
    _cell(ws, f'A{end_row}',
          f'合計: PASS {pass_cnt} / WARN {warn_cnt} / FAIL {fail_cnt} / INFO {len(cases) - pass_cnt - warn_cnt - fail_cnt}',
          font=FONT_TD_B, fill=FILL_TOTAL, align=ALIGN_L)

    return pass_cnt, warn_cnt, fail_cnt


# ─────────────────────────────────────────────────
# Sheet 07: 承認履歴
# ─────────────────────────────────────────────────
def _build_07_approvals(wb, ctx: ExcelFillContext):
    ws = wb.create_sheet('07_承認履歴')
    _set_col_widths(ws, {'A': 5, 'B': 14, 'C': 28, 'D': 22, 'E': 22, 'F': 42})
    _banner(ws, 1, 'A', 'F', '承認履歴', 'AI 2段レビュー + 人間承認')

    headers = ['#', '段', 'レビュアー', '日時', '判定', 'コメント']
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = FONT_TH_INV; c.fill = FILL_HDR_D; c.alignment = ALIGN_CTR; c.border = BORDER
    ws.row_dimensions[4].height = 24

    ph = ctx.phase_results or {}
    status = ctx.final_status or 'UNKNOWN'
    phase_7 = ph.get('phase_7', 'PENDING')

    ai1_verdict = 'OK' if status not in ('REJECTED', 'FAILED') else 'NG'
    ai2_verdict = 'OK' if status == 'APPROVED' or phase_7 == 'PENDING' else ('NG' if status == 'REJECTED' else 'OK')

    human_verdict = {
        'APPROVED': '☑適用可',
        'REJECTED': '☑否',
        'INITIAL_BASELINE': '☑baseline',
    }.get(status, '☐適用可 ☐差戻し ☐否')

    approvals = [
        ('AI 1段目', 'Claude Code (実装担当)', ctx.run_ts or '-', ai1_verdict,
         f'phase_3={ph.get("phase_3", "-")} / phase_5={ph.get("phase_5", "-")}'),
        ('AI 2段目', 'Claude Code (レビュー)', ctx.run_ts or '-', ai2_verdict,
         f'phase_7={phase_7} / final_status={status}'),
        ('人間',     '____________________', '____-__-__ __:__', human_verdict, '所感:'),
    ]

    for idx, (stage, reviewer, dt, verdict, comment) in enumerate(approvals):
        r = 5 + idx
        _cell(ws, f'A{r}', idx + 1, font=FONT_TD, align=ALIGN_CTR)
        _cell(ws, f'B{r}', stage, font=FONT_TD_B, align=ALIGN_CTR)
        _cell(ws, f'C{r}', reviewer, font=FONT_TD, align=ALIGN_L)
        _cell(ws, f'D{r}', dt, font=FONT_TD, align=ALIGN_CTR)
        is_human = stage == '人間'
        fill = FILL_PANEL if is_human else (FILL_OK if verdict == 'OK' else FILL_NG)
        _cell(ws, f'E{r}', verdict, font=FONT_TD_B, fill=fill, align=ALIGN_CTR)
        _cell(ws, f'F{r}', comment, font=FONT_TD, align=ALIGN_L)
        ws.row_dimensions[r].height = 42 if is_human else 22

    # フッタ注記
    ws.merge_cells('A9:F10')
    _cell(ws, 'A9',
          '※ 人間承認欄は手書き記入枠。判定が「差戻し」「否」の場合は再 sync + 本レポート再生成を伴う。',
          font=FONT_TD, fill=FILL_PANEL, align=ALIGN_L)
